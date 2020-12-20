import csv
from openpyxl import load_workbook

county = "Winston"

wb = load_workbook(filename = f"/Users/derekwillis/Downloads/{county}_REP_DEM.xlsx")

results = []

for sheet in wb.worksheets:
    if 'REP' in sheet.title:
        party = 'REP'
    else:
        party = 'DEM'
    precincts = [sheet.cell(1, x).value for x in range(1, sheet.max_column+1) if sheet.cell(1, x).value is not None]
    candidates_with_offices = [sheet.cell(x, 1).value for x in range(1, sheet.max_row+1) if sheet.cell(x, 1).value is not None]
    for index, c_or_o in enumerate(candidates_with_offices):
        if c_or_o == 'United States-President':
            candidate = None
            office = 'President'
            district = None
        elif c_or_o == 'United States-Senate':
            candidate = None
            office = 'U.S. Senate'
            district = None
        elif 'US House Of Rep' in c_or_o:
            candidate = None
            office = 'U.S. House'
            district = c_or_o.split('-')[1][0]
        else:
            candidate = c_or_o
            vote_cols = [sheet.cell(index+2, x).value for x in range(3, sheet.max_column+1) if sheet.cell(index+2, x).value is not None]
            print(candidate)
            print(vote_cols)
            precincts_with_votes = zip(precincts, vote_cols)
            for precinct, votes in precincts_with_votes:
                results.append([county, precinct, office, district, candidate, votes])

with open(f"20200310__ms__primary__{county.lower().replace(' ','_')}__precinct.csv", 'w') as output_file:
    outfile = csv.writer(output_file)
    outfile.writerow(['county','precinct', 'office', 'district', 'candidate', 'party', 'votes'])
    outfile.writerows(results)
